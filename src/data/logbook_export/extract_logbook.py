"""Extract IC_Logbook.xlsx into clean, DB-ready CSVs.

Outputs (under src/data/logbook_export/):
  - ic_logbook.csv          : all experiments, one row each, with a `category`
                              column + the standard fields + the extra Sulfidized
                              synthesis columns (empty for other categories).
  - by_category/<cat>.csv   : the same, split one file per category.
  - summary_scoreboard.csv  : the manual scoreboard snapshots from the Summary sheet.
"""
import csv
import os
import re
from datetime import datetime, date

import openpyxl

SRC = "src/data/IC_Logbook.xlsx"
OUT = "src/data/logbook_export"
BY_CAT = os.path.join(OUT, "by_category")

# Ordered output columns for the combined/per-category CSVs.
FIELDS = [
    "category", "ic_id", "owner", "assembled_by", "disassembled_by",
    "test_length_h", "protocol", "format", "experiment_finished_on",
    "experiment_finished_on_raw", "notes", "archived_in", "plot",
    "synthesis_recipe", "synthesis_time_h", "synthesis_temperature_c",
    "coating_test_on",
]

# Known source-data corrections applied during extraction (see
# ic_logbook_processing.md). The Excel still holds the mistaken values (three
# Test-length cells typed as h:mm:ss times), so fixing them here keeps every
# re-extraction correct. (category, ic_id) -> {field: corrected_value}.
CORRECTIONS = {
    ("Gen2-Gen2", "IC0570_CatNNF_AnoNNF_rho1276_AUX1Cat_AUX2Ano_3-8"): {"test_length_h": "168"},
    ("Gen2-Gen2", "IC0571_CatNNF_AnoNNF_rho1276_AUX1Cat_AUX2Ano_3-7"): {"test_length_h": "168"},
    ("Gen3Cat-Gen2", "IC1355_CatB8111-TR-P2_AnoNNF_rho1281_4-4"): {"test_length_h": "48"},
}


def field_for(header) -> str | None:
    """Map a (messy) sheet header to one of our field names, or None."""
    h = re.sub(r"\s+", " ", str(header or "").strip().lower())
    if not h:
        return None
    if h.startswith("ic_id") or h.startswith("ic id"):
        return "ic_id"
    if h == "owner":
        return "owner"
    if h.startswith("assembled"):
        return "assembled_by"
    if h.startswith("disassembled"):
        return "disassembled_by"
    if h.startswith("test length"):
        return "test_length_h"
    if h == "protocol":
        return "protocol"
    if h == "format":
        return "format"
    if h.startswith("experiment finished"):
        return "experiment_finished_on"
    if h == "notes":
        return "notes"
    if h.startswith("archived") or h.startswith("arhived"):
        return "archived_in"
    if h == "plot":
        return "plot"
    if h.startswith("synthesis recipe"):
        return "synthesis_recipe"
    if h.startswith("synthesis time"):
        return "synthesis_time_h"
    if h.startswith("synthesis temp"):
        return "synthesis_temperature_c"
    if h.startswith("coating test"):
        return "coating_test_on"
    return None


unparseable_dates = {}


def norm_date(v):
    """Return (iso_or_blank, raw_string). Dates are European (dd/mm/yyyy)."""
    if v is None:
        return "", ""
    if isinstance(v, datetime):
        return v.date().isoformat(), v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat(), v.isoformat()
    s = str(v).strip()
    if not s:
        return "", ""
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat(), s
        except ValueError:
            continue
    unparseable_dates[s] = unparseable_dates.get(s, 0) + 1
    return "", s  # keep the raw so nothing is lost


def cell(v):
    """Stringify a cell value for CSV (numbers/dates preserved sensibly)."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def main():
    os.makedirs(BY_CAT, exist_ok=True)
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    all_rows = []
    per_cat_counts = {}

    for name in wb.sheetnames:
        if name.strip().lower().startswith("1. summary") or name.strip().lower() == "summary":
            continue
        category = re.sub(r"\s+", " ", name.strip())
        ws = wb[name]

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            continue
        col_field = {i: field_for(h) for i, h in enumerate(header)}

        blank_streak = 0
        cat_rows = []
        for row in rows_iter:
            rec = {f: "" for f in FIELDS}
            rec["category"] = category
            has_data = False  # any mapped source field populated (not just ic_id)
            for i, val in enumerate(row):
                f = col_field.get(i)
                if not f:
                    continue
                if f == "experiment_finished_on":
                    iso, raw = norm_date(val)
                    rec["experiment_finished_on"] = iso
                    rec["experiment_finished_on_raw"] = raw
                    if raw:
                        has_data = True
                else:
                    rec[f] = cell(val)
                    if rec[f]:
                        has_data = True
            # Keep any row with real data even if the IC_ID is missing (those are
            # still experiments; the DB's auto-increment id is the key). Only a
            # fully-empty row counts toward the end-of-data streak.
            if not has_data:
                blank_streak += 1
                if blank_streak > 100:  # data has ended (oversized sheets)
                    break
                continue
            blank_streak = 0
            fix = CORRECTIONS.get((rec["category"], rec["ic_id"]))
            if fix:
                rec.update(fix)
            cat_rows.append(rec)

        all_rows.extend(cat_rows)
        per_cat_counts[category] = len(cat_rows)

        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", category).strip("_")
        with open(os.path.join(BY_CAT, f"{safe}.csv"), "w", newline="",
                  encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=FIELDS)
            w.writeheader()
            w.writerows(cat_rows)

    # Combined file.
    with open(os.path.join(OUT, "ic_logbook.csv"), "w", newline="",
              encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(all_rows)

    # Scoreboard from the Summary sheet (manual historical snapshots).
    ws = wb["1. Summary"]
    sb = []
    grab = False
    for row in ws.iter_rows(values_only=True):
        a = (str(row[0]).strip().lower() if row[0] is not None else "")
        if a == "scoreboard":
            grab = True
            continue
        if grab:
            if row[0] is None:
                continue
            if a == "date":  # header row
                continue
            sb.append([cell(row[0]), cell(row[1]), cell(row[2]), cell(row[3])])
    with open(os.path.join(OUT, "summary_scoreboard.csv"), "w", newline="",
              encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["date", "days_since_id_intro", "total_testing_time_h",
                    "number_of_ics"])
        w.writerows(sb)

    # Report.
    print(f"Wrote {len(all_rows)} experiments to {OUT}/ic_logbook.csv")
    print(f"Per-category files in {BY_CAT}/")
    for c, n in per_cat_counts.items():
        print(f"  {c:<22} {n}")
    print(f"Scoreboard rows: {len(sb)}")
    if unparseable_dates:
        print("\nUnparseable 'Experiment finished on' values (kept in _raw):")
        for s, n in sorted(unparseable_dates.items(), key=lambda x: -x[1]):
            print(f"  {n:>4}x  {s!r}")
    else:
        print("\nAll 'Experiment finished on' values parsed to ISO dates.")


if __name__ == "__main__":
    main()
